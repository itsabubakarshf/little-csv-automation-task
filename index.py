import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font

def extract_data_from_file(file_path):
    identifier = os.path.basename(file_path).split('_')[0]
    
    df = pd.read_csv(file_path, skiprows=9)
    
    values = df['Total Cashflow'].dropna().values[1:]  # Skipping the "Total Cashflow" header
    
    return identifier, values

def gather_data_from_all_files(directory):
    final_df = pd.DataFrame()
    
    for file_name in os.listdir(directory):
        if file_name.startswith("EPAAG6") and file_name.endswith(".csv"):
            identifier, values = extract_data_from_file(os.path.join(directory, file_name))
            final_df[identifier] = pd.Series(values)
    
    final_df['Total'] = final_df.sum(axis=1)
    
    return final_df                                         

input_directory_path = "Files/AG38_V6_S130/"  # Replace with your directory path where the files are located

output_directory_path = "Files"  # Replace with your desired output directory path

final_data = gather_data_from_all_files(input_directory_path)

output_path = os.path.join(output_directory_path, "consolidated_data.xlsx")
final_data.to_excel(output_path, index=False)


summary_wb = openpyxl.load_workbook('Files/summary.xlsx')
consolidated_wb = openpyxl.load_workbook('Files/consolidated_data.xlsx')

summary_ws = summary_wb['Sheet1']
consolidated_ws = consolidated_wb.active

data_to_copy = list(zip(summary_ws['A'][2:], summary_ws['B'][2:], summary_ws['C'][2:]))

original_data = []
for row in consolidated_ws.iter_rows():
    original_data.append([(cell.value, cell.number_format) for cell in row])

for row in consolidated_ws.iter_rows():
    for cell in row:
        cell.value = None

for idx, (a, b, c) in enumerate(data_to_copy):
    consolidated_ws.cell(row=idx + 1, column=1, value=a.value)
    consolidated_ws.cell(row=idx + 1, column=2, value=b.value)
    consolidated_ws.cell(row=idx + 1, column=3, value=c.value)

for row_idx, row in enumerate(original_data, start=1):
    for col_idx, (value, number_format) in enumerate(row, start=4):
        cell = consolidated_ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.number_format = number_format

bold_font = Font(bold=True)
last_col = consolidated_ws.max_column
consolidated_ws.cell(row=1, column=last_col - 2).font = bold_font  # For EPAAG6AA
consolidated_ws.cell(row=1, column=last_col - 1).font = bold_font  # For EPAAG6CE
consolidated_ws.cell(row=1, column=last_col).font = bold_font      # For Total

data_to_shift = []
for row in consolidated_ws.iter_rows(min_row=2, min_col=4):
    data_to_shift.append([(cell.value, cell.number_format) for cell in row])

for row in consolidated_ws.iter_rows(min_row=2, min_col=4):
    for cell in row:
        cell.value = None

for row_idx, row_data in enumerate(data_to_shift, start=3):
    for col_idx, (value, number_format) in enumerate(row_data, start=4):
        cell = consolidated_ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.number_format = number_format

consolidated_wb.save('final_output_consolidated_data.xlsx')
output_wb = openpyxl.load_workbook('final_output_consolidated_data.xlsx')
output_ws = output_wb.active
for row in output_ws.iter_rows():
    for cell in row:
        if cell.value == 0 or cell.value is None:
            cell.value = "-"
output_wb.save('final_output_consolidated_data.xlsx')
print(f"Data saved to: {output_path}")
